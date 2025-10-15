# app.py
import os
import io
import re
import json
import time
import base64
import unicodedata
from pathlib import Path
from datetime import datetime

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# =============== Config página / estilo ===============
st.set_page_config(page_title="Carga SIOT", page_icon="📤", layout="wide")

# ---- Estilos (Arial + botón naranja + uploader beige) ----
st.markdown("""
<style>
html, body, [class*="css"] { font-family: 'Arial', sans-serif !important; }
.block-container { max-width: 1200px; padding-top: 3.25rem !important; margin-top: 0 !important; }
h1, h2, h3 { font-weight: 800; }
.kpi { padding: 12px 16px; border-radius: 14px; background: #fff; box-shadow: 0 3px 12px rgba(0,0,0,.06); border: 1px solid #eee; }

/* ============ Botones ============ */
div.stButton > button {
  background: #FF7A00 !important;
  border: 1px solid #FF7A00 !important;
  color: #ffffff !important;
  font-weight: 700 !important;
  border-radius: 12px !important;
  padding: 0.6rem 1rem !important;
  box-shadow: 0 4px 10px rgba(255,122,0,0.25) !important;
}
div.stButton > button:hover { background: #E56D00 !important; border-color: #E56D00 !important; }
div.stButton > button:focus, div.stButton > button:active {
  background: #CC6000 !important; border-color: #CC6000 !important;
  box-shadow: 0 0 0 3px rgba(255,122,0,0.25) !important;
}
/* Forzar naranja si kind="primary" */
button[kind="primary"]{ background:#FF7A00 !important; border:1px solid #FF7A00 !important; color:#fff !important; }
button[kind="primary"]:hover{ background:#E56D00 !important; border-color:#E56D00 !important; }

/* ============ File Uploader en beige ============ */
[data-testid="stFileUploaderDropzone"],
div[aria-label="Upload area"]{
  background: #F6EFE6 !important;
  border: 1.5px dashed #E3D5C3 !important;
  border-radius: 14px !important;
}
[data-testid="stFileUploaderDropzone"] * { color: #4A3F33 !important; }
[data-testid="stFileUploaderDropzone"]:hover{ background:#F2E8DB !important; border-color:#D9C7B2 !important; }
[data-testid="stFileUploaderDropzone"] svg { fill:#CC6000 !important; }
[data-testid="stFileUploader"] .uploadedFile {
  background:#F6EFE6 !important; border:1px solid #E3D5C3 !important; color:#4A3F33 !important;
}
</style>
""", unsafe_allow_html=True)

# =============== Logo ===============
def _find_logo_bytes() -> bytes | None:
    for p in [
        Path("/mnt/data/06ccb9c2-ca99-49b6-a58e-9452a7e6a452.png"),
        Path("/mnt/data/Logo EOMMT.png"),
        Path(__file__).parent / "Logo EOMMT.png",
        Path("Logo EOMMT.png"),
        Path("logo_eommt.png"),
    ]:
        try:
            if p.exists():
                return p.read_bytes()
        except Exception:
            pass
    return None

def render_logo_center(width_px: int = 220):
    img = _find_logo_bytes()
    if not img: return
    b64 = base64.b64encode(img).decode("ascii")
    st.markdown(
        f'<div style="text-align:center;margin:6px 0 10px 0;"><img src="data:image/png;base64,{b64}" width="{width_px}"/></div>',
        unsafe_allow_html=True
    )

def render_logo_sidebar(width_px: int = 160):
    img = _find_logo_bytes()
    if not img: return
    b64 = base64.b64encode(img).decode("ascii")
    st.sidebar.markdown(
        f'<div style="text-align:center;margin:6px 0 10px 0;"><img src="data:image/png;base64,{b64}" width="{width_px}"/></div>',
        unsafe_allow_html=True
    )

# =============== Auth simple ===============
AUTH_USERS = json.loads(os.environ.get("AUTH_USERS_JSON", os.getenv("AUTH_USERS_JSON", '{"admin":"admin"}')))

def login_view():
    _, c, _ = st.columns([1,1,1])
    with c:
        render_logo_center(200)
        st.markdown("## 🚇 Instrucción Operacional de Trabajos")
        st.markdown("### Ingreso al sistema")
        user = st.text_input("Usuario", placeholder="Escribe tu usuario")
        pwd  = st.text_input("Contraseña", type="password", placeholder="Escribe tu contraseña")
        if st.button("Ingresar", use_container_width=True):
            if user in AUTH_USERS and AUTH_USERS.get(user) == pwd:
                st.session_state["auth_user"] = user
                st.success("✅ Acceso concedido.")
                st.rerun()
            else:
                st.error("❌ Usuario o contraseña incorrectos.")
    return "auth_user" in st.session_state

def require_auth():
    if "auth_user" in st.session_state: return True
    ok = login_view()
    if not ok: st.stop()
    return True

def logout_button():
    with st.sidebar:
        if st.button("🚪 Cerrar sesión", use_container_width=True):
            st.session_state.pop("auth_user", None)
            st.rerun()

# =============== Secrets / Pipefy ===============
PIPEFY_API_URL = "https://api.pipefy.com/graphql"
def get_secret(name, default=None):
    try: return st.secrets[name]
    except Exception: return os.getenv(name, default)

PIPEFY_TOKEN = get_secret("PIPEFY_TOKEN", "")
PIPE_ID      = int(str(get_secret("PIPEFY_PIPE_ID", "0")) or "0")

# =============== Normalización de columnas ===============
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def _normalize_key(s: str) -> str:
    """Normaliza encabezados: quita paréntesis, saltos de línea, asteriscos y tildes."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\n", " ")                 # quita salto de línea
    s = re.sub(r"\(.*?\)", "", s)            # quita '(...)'
    s = s.replace("*", " ")                  # quita asteriscos
    s = _strip_accents(s)                    # sin tildes
    s = re.sub(r"[^A-Z0-9/ ]", " ", s.upper())  # deja letras/números/espacio y '/'
    s = re.sub(r"\s+", " ", s).strip()       # colapsa espacios
    return s

# -------- Mapa de alias -> nombre canónico (los que usa el envío a Pipefy) --------
ALIASES = {
    "EMPRESA": ["EMPRESA"],

    "CCU": [
        "CCU",
        "CCU COORDINADOR DE CUADRILLA NOMBRE APELLIDO",
        "COORDINADOR DE CUADRILLA",
        "NOMBRE COORDINADOR",
        "CCU NOMBRE APELLIDO",
    ],

    "INTEGRANTES DE CUADRILLA": [
        "INTEGRANTES DE CUADRILLA",
        "INTEGRANTES DEL EQUIPO DE CUADRILLA",
        "INTEGRANTES DEL EQUIPO DE CUADRILLA NOMBRE APELLIDO - NUMERO DE CEDULA",
    ],

    "CONTACTO CCU": [
        "CONTACTO CCU",
        "TELEFONO DE CONTACTO CCU",
        "TELEFONO COORDINADOR",
    ],

    # ← canónico para email
    "CORREO DEL SOLICITANTE": [
        "CORREO DEL SOLICITANTE",
        "CORREO ELECTRONICO DEL SOLICITANTE",
        "EMAIL", "E MAIL", "CORREO ELECTRONICO",
    ],

    "FECHA DE INICIO": ["FECHA DE INICIO", "FECHA INICIO", "INICIO FECHA"],
    "FECHA DE FIN":    ["FECHA DE FIN", "FECHA FIN", "FIN FECHA"],
    "HORA DE INICIO":  ["HORA DE INICIO", "HORA INICIO", "INICIO HORA"],
    "HORA DE FIN":     ["HORA DE FIN", "HORA FIN", "FIN HORA"],

    "CANTÓN / ESTACIÓN": ["CANTON / ESTACION", "CANTON", "ESTACION", "CANTON / ESTACION"],

    # plural → canónico singular
    "ZONA DE ESTACIÓN": ["ZONA DE ESTACION", "ZONAS DE ESTACION", "ZONAS DE ESTACION "],

    "CATEGORÍA DE TRABAJOS": ["CATEGORIA DE TRABAJOS"],
    "TIPO DE MANTENIMIENTO / INSPECCIÓN": [
        "TIPO DE MANTENIMIENTO / INSPECCION",
        "TIPO DE MANTENIMIENTO", "TIPO DE INSPECCION",
    ],

    # con y sin "DE"
    "N° REGISTRO DE FALLA": ["N° REGISTRO FALLA", "N REGISTRO FALLA", "NUMERO REGISTRO FALLA"],

    "CATEGORÍA DE RIESGO": ["CATEGORIA DE RIESGO"],
    "DESCRIPCIÓN DE ACTIVIDAD": ["DESCRIPCION DE ACTIVIDAD", "DESCRIPCION", "ACTIVIDAD"],
    "DESENERGIZACIONES": ["DESENERGIZACIONES", "DESENERGIZACION"],

    # con asterisco o “DE LA ZONA”
    "VEHÍCULO": ["VEHICULO", "VEHICULOS", "VEHICULO "],
    "ILUMINACIÓN PARCIAL": ["ILUMINACION PARCIAL", "ILUMINACION PARCIAL DE LA ZONA"],
    "SEÑALETICA PROPIA": ["SENALETICA PROPIA", "SENHALETICA PROPIA", "SEÑALÉTICA PROPIA"],

    "R1": ["R1"], "R2": ["R2"], "P1": ["P1"], "P3": ["P3"], "E1": ["E1"], "V3": ["V3"],
    "P6": ["P6"], "P7": ["P7"], "P8": ["P8"],

    "BLOQUEO DE VÍA": ["BLOQUEO DE VIA", "BLOQUEO DE VÍA"],
    "DESDE": ["DESDE"],
    "HASTA": ["HASTA"],

    "Seleccionar etiqueta": ["SELECCIONAR ETIQUETA", "ETIQUETA"],
}

def apply_aliases(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    norm_cols = {_normalize_key(c): c for c in df.columns}
    rename_map = {}
    for canon, variants in ALIASES.items():
        for v in (variants + [canon]):  # intentamos variantes + el canónico
            vn = _normalize_key(v)
            if vn in norm_cols:
                rename_map[norm_cols[vn]] = canon
                break
    if rename_map:
        df = df.rename(columns=rename_map)
    return df

# =============== Utilidades de Excel ===============
def read_excel_table_siot(uploaded_bytes: bytes, table_name: str = "SIOT") -> pd.DataFrame:
    """Lee la tabla SIOT; si no existe, fallback por encabezado flexible."""
    bio = io.BytesIO(uploaded_bytes)
    wb = load_workbook(bio, data_only=True, read_only=False)

    # 1) Intentar tabla SIOT
    for ws in wb.worksheets:
        tables = ws.tables or {}
        for t in tables.values():
            if (t.name or "").strip().lower() == table_name.lower():
                start, end = t.ref.split(":")
                start_col = ''.join(filter(str.isalpha, start))
                start_row = int(''.join(filter(str.isdigit, start)))
                end_col = ''.join(filter(str.isalpha, end))
                end_row = int(''.join(filter(str.isdigit, end)))
                min_col = column_index_from_string(start_col)
                max_col = column_index_from_string(end_col)
                data = []
                for r in ws.iter_rows(min_row=start_row, max_row=end_row,
                                      min_col=min_col, max_col=max_col, values_only=True):
                    data.append(list(r))
                if not data: return pd.DataFrame()
                header = [str(h).strip() if h is not None else "" for h in data[0]]
                body = data[1:]
                df = pd.DataFrame(body, columns=header)
                df = df.loc[:, [c for c in df.columns if str(c).strip() and not str(c).startswith("Unnamed")]]
                for c in df.columns:
                    if df[c].dtype == object:
                        df[c] = df[c].apply(lambda x: x.strip() if isinstance(x, str) else x)
                return df.dropna(how="all")

    # 2) Fallback: detectar encabezado por EMPRESA (o similares)
    bio.seek(0)
    raw = pd.read_excel(bio, engine="openpyxl", sheet_name=0, header=None)
    header_row = None
    for i in range(min(60, len(raw))):
        vals = [str(x).strip().upper() if pd.notna(x) else "" for x in raw.iloc[i].tolist()]
        # si alguna celda parece EMPRESA (según alias)
        if any(_normalize_key(v) in [_normalize_key(a) for a in ALIASES.get("EMPRESA", ["EMPRESA"])] for v in vals):
            header_row = i
            break
    if header_row is None: return pd.DataFrame()
    headers = [str(c).strip() if pd.notna(c) else "" for c in raw.iloc[header_row].tolist()]
    df = raw.iloc[header_row+1:].copy()
    df.columns = headers
    df = df.loc[:, [c for c in df.columns if c and not str(c).startswith("Unnamed")]]
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].apply(lambda x: x.strip() if isinstance(x, str) else x)
    return df.dropna(how="all")

# =============== Utilidades de campos/fechas/labels ===============
def _fmt_date(val):
    if val is None: return None
    try:
        if isinstance(val, float) and pd.isna(val): return None
    except Exception: pass
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() == "nan": return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception: pass
    return s

def _add_field(fields, field_id, value):
    if value is None: return
    try:
        if isinstance(value, float) and pd.isna(value): return
    except Exception: pass
    s = str(value).strip()
    if not s or s.lower() == "nan": return
    fields.append({"field_id": field_id, "field_value": s})

def _parse_multi(val):
    if val is None: return None
    try:
        if isinstance(val, float) and pd.isna(val): return None
    except Exception: pass
    if isinstance(val, (list, tuple, set)):
        out = [str(x).strip() for x in val if str(x).strip() not in ("", "nan")]
        return out or None
    s = str(val).strip()
    if not s or s.lower() == "nan": return None
    return [p.strip() for p in s.replace(",", ";").split(";") if p.strip()] or None

def _add_field_list(fields, field_id, value):
    items = _parse_multi(value)
    if items:
        fields.append({"field_id": field_id, "field_value": items})

def _fetch_labels_map(token: str, pipe_id: int) -> dict:
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    q = {"query": "query($id: ID!){ pipe(id:$id){ labels{ id name } } }", "variables": {"id": pipe_id}}
    try:
        r = requests.post(PIPEFY_API_URL, headers=headers, json=q, timeout=30)
        if r.status_code != 200: return {}
        data = r.json().get("data", {}).get("pipe", {}).get("labels", [])
        return {x["name"]: x["id"] for x in data if "id" in x and "name" in x}
    except Exception:
        return {}

def _add_label_select(fields, field_id, value, labels_map, report_missing):
    items = _parse_multi(value)
    if not items: return
    ids = []
    for name in items:
        name = str(name).strip()
        if not name: continue
        if name in labels_map: ids.append(labels_map[name])
        else: report_missing.append(name)
    if ids:
        fields.append({"field_id": field_id, "field_value": ids})

def pipefy_create_card(token: str, pipe_id: int, fields_attrs: list, title: str):
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    mutation = {
        "query": """
        mutation($input: CreateCardInput!) {
          createCard(input: $input) { card { id title } }
        }
        """,
        "variables": {"input": {"pipe_id": pipe_id, "title": title, "fields_attributes": fields_attrs}},
    }
    try:
        resp = requests.post(PIPEFY_API_URL, headers=headers, json=mutation, timeout=40)
        if resp.status_code != 200:
            return False, f"HTTP {resp.status_code}: {resp.text}"
        data = resp.json()
        if "errors" in data: return False, str(data["errors"])
        return True, data.get("data", {}).get("createCard", {}).get("card", {}).get("id")
    except Exception as e:
        return False, str(e)

# =============== Reglas de obligatoriedad ===============
REQUIRED_COLS = [
    "CCU",
    "INTEGRANTES DE CUADRILLA",
    "CONTACTO CCU",
    "ZONA DE ESTACIÓN",
    "FECHA DE INICIO",
    "FECHA DE FIN",
    "HORA DE INICIO",
    "HORA DE FIN",
    "N° REGISTRO DE FALLA",
    "VEHÍCULO",
    "ILUMINACIÓN PARCIAL",
    "SEÑALETICA PROPIA",
    "CORREO DEL SOLICITANTE",
]

# =============== APP ===============
if require_auth():
    render_logo_sidebar(150)
    logout_button()

    render_logo_center(220)
    st.title("INSTRUCCIÓN OPERACIONAL DE TRABAJOS")

    if not PIPEFY_TOKEN or not PIPE_ID:
        st.error("Faltan credenciales en `st.secrets`: agrega `PIPEFY_TOKEN` y `PIPEFY_PIPE_ID`.")
        st.stop()

    up = st.file_uploader("Sube tu Excel (.xlsx) con la tabla **SIOT**", type=["xlsx"])

    if up is not None:
        content = up.read()
        df = read_excel_table_siot(content, "SIOT")

        if df.empty:
            st.error("No se logró leer datos de la tabla **SIOT** ni por fallback de encabezados.")
            st.stop()

        # Renombrar columnas usando alias/normalización
        df = apply_aliases(df)

        # Cortar hasta última fila con EMPRESA no vacía (si existe la columna)
        if "EMPRESA" in df.columns:
            mask_emp = df["EMPRESA"].astype(str).str.strip().replace({"None": "", "nan": ""}) != ""
            if mask_emp.any():
                df = df.loc[df.index.min(): df.index[mask_emp].max()].copy()

        st.subheader("👀 Vista previa")
        st.dataframe(df.head(50), use_container_width=True)

        # === Validación de obligatorios (por fila) ===
        faltantes_por_fila = []
        for idx, row in df.iterrows():
            faltan = [c for c in REQUIRED_COLS if c in df.columns and (pd.isna(row.get(c)) or str(row.get(c)).strip() == "")]
            if faltan:
                faltantes_por_fila.append({"fila": int(idx)+1, "faltan": ", ".join(faltan)})

        valid_mask = pd.Series(True, index=df.index)
        for item in faltantes_por_fila:
            i = item["fila"] - 1
            if i in valid_mask.index: valid_mask.loc[i] = False

        df_validas = df[valid_mask].copy()
        df_invalidas = df[~valid_mask].copy()

        c1, c2, c3 = st.columns(3)
        with c1: st.markdown(f"<div class='kpi'><b>Filas totales</b><br>{len(df)}</div>", unsafe_allow_html=True)
        with c2: st.markdown(f"<div class='kpi'><b>Filas válidas</b><br>{len(df_validas)}</div>", unsafe_allow_html=True)
        with c3: st.markdown(f"<div class='kpi'><b>Filas con faltantes</b><br>{len(df_invalidas)}</div>", unsafe_allow_html=True)

        if df_invalidas.shape[0] > 0:
            st.warning("Hay filas con **campos obligatorios** vacíos. No se subirán. Detalle:")
            st.dataframe(pd.DataFrame(faltantes_por_fila), use_container_width=True)

        # ===== Botón para subir SOLO filas válidas =====
        if st.button(f"🚀 Subir a Pipefy ({len(df_validas)} tarjetas)", type="primary", use_container_width=True, disabled=(len(df_validas) == 0)):
            labels_map = _fetch_labels_map(PIPEFY_TOKEN, PIPE_ID)
            creadas = errores = 0
            missing_labels = []

            progress = st.progress(0.0, text="Iniciando…")
            total = len(df_validas)

            for i, (_, row) in enumerate(df_validas.iterrows(), start=1):
                fields = []
                # Texto/fecha/select
                _add_field(fields, "empresa", row.get("EMPRESA"))
                _add_field(fields, "ccu_1", row.get("CCU"))
                _add_field(fields, "integrantes_de_cuadrilla", row.get("INTEGRANTES DE CUADRILLA"))
                _add_field(fields, "contacto_coordinador_de_cuadrilla", row.get("CONTACTO CCU"))
                _add_field(fields, "fecha_de_inicio", _fmt_date(row.get("FECHA DE INICIO")))
                _add_field(fields, "fecha_de_fin", _fmt_date(row.get("FECHA DE FIN")))
                _add_field(fields, "cant_n_estaci_n", row.get("CANTÓN / ESTACIÓN"))
                _add_field(fields, "zona_de_trabajo", row.get("ZONA DE ESTACIÓN"))
                _add_field(fields, "descripci_n_de_actividad", row.get("DESCRIPCIÓN DE ACTIVIDAD"))
                _add_field(fields, "hora_de_inicio", row.get("HORA DE INICIO"))
                _add_field(fields, "hora_de_fin", row.get("HORA DE FIN"))
                _add_field(fields, "tipo_de_jornada", row.get("TIPO DE JORNADA"))
                _add_field(fields, "tipo_de_mantenimiento", row.get("TIPO DE MANTENIMIENTO / INSPECCIÓN"))
                _add_field(fields, "registro_de_incidente", row.get("N° REGISTRO DE FALLA"))
                _add_field(fields, "categor_a_de_riesgo", row.get("CATEGORÍA DE RIESGO"))
                _add_field(fields, "categor_a_de_trabajos", row.get("CATEGORÍA DE TRABAJOS"))
                _add_field(fields, "desenergizaci_n", row.get("DESENERGIZACIONES"))

                # Checklists / multiselect
                _add_field_list(fields, "veh_culo", row.get("VEHÍCULO"))
                _add_field_list(fields, "iluminaci_n_parcia", row.get("ILUMINACIÓN PARCIAL"))
                _add_field_list(fields, "se_aletica_propia", row.get("SEÑALETICA PROPIA"))
                _add_field_list(fields, "r1_1", row.get("R1"))
                _add_field_list(fields, "r2_1", row.get("R2"))
                _add_field_list(fields, "p1", row.get("P1"))
                _add_field_list(fields, "p3", row.get("P3"))
                _add_field_list(fields, "e1", row.get("E1"))
                _add_field_list(fields, "v3", row.get("V3"))
                _add_field_list(fields, "copy_of_se_aletica_propia", row.get("P6"))
                _add_field_list(fields, "copy_of_r1", row.get("P7"))
                _add_field_list(fields, "copy_of_p3", row.get("P8"))

                # Bloqueo de vía
                _add_field_list(fields, "bloqueo_de_v_a_1", row.get("BLOQUEO DE VÍA"))
                _add_field(fields, "bloqueo_desde", row.get("DESDE"))
                _add_field(fields, "hasta", row.get("HASTA"))

                # Etiquetas (label_select)
                _add_label_select(fields, "seleccionar_etiqueta", row.get("Seleccionar etiqueta"), labels_map, missing_labels)

                # Email solicitante (canónico que definimos)
                _add_field(fields, "correo_electr_nico_del_solicitante", row.get("CORREO DEL SOLICITANTE"))

                title = str(row.get("EMPRESA") or f"Fila {i}")
                ok, info = pipefy_create_card(PIPEFY_TOKEN, PIPE_ID, fields, title)
                if ok:
                    creadas += 1
                else:
                    errores += 1
                    st.error(f"❌ Error en fila {i}: {info}")

                time.sleep(0.05)
                progress.progress(i/total, text=f"Procesadas {i}/{total}")

            if missing_labels:
                st.warning("Estas etiquetas NO existen en el Pipe y se omitieron: " + ", ".join(sorted(set(missing_labels))))
            st.success(f"✅ Terminado. Tarjetas creadas: {creadas} • Errores: {errores}")


